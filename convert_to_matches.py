# This files reads the file WordFeud.xlsx and gets all matches from the sheets
# and writes them to the file matches.txt

import openpyxl
import time
import logging

#logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)

def is_number(s):
    return isinstance(s, int) or isinstance(s, float)

def is_season_score(player1, player2, status, score1, score2):
    try:
        return player1.strip() != "" and player2.strip() != "" and status == "Avslutad" and is_number(score1) and is_number(score2)
    except Exception:
        return False

def is_cup_score(player1, player2, score1, score2, empty):
    try:
        return player1.strip() != "" and player2.strip() != "" and is_number(score1) and is_number(score2) and empty is None
    except Exception:
        return False

def is_alternate_cup_score(player1, player2, status, score1, score2):
    try:
        return player1.strip() != "" and player2.strip() != "" and status == "Avslutad" and is_number(score1) and is_number(score2)
    except Exception:
        return False

spreadsheet_file = "WordFeud.xlsx"
rayter_file = "wordfeudligan.txt"

with open(rayter_file, "w") as f:
    f.write("game_name Wordfeudligan\n")

    # Open the file
    print("Loading " + spreadsheet_file)
    workbook = openpyxl.load_workbook(spreadsheet_file)

    matches = []

    wordfeudname_to_realname = {
        "äggeth": "PelleÅ",
        "Äggeth": "PelleÅ",
        "Andreas112": "Andreas",
        "Annafian": "Fia",
        "annafian": "Fia",
        "CamillaB71": "Camilla",
        "dinkelidunk": "Tommy",
        "Doktorn72": "David",
        "drsadness": "Jonas",
        "Emma-Victoria": "Emma",
        "gurka1495": "Per",
        "johlits": "JohanL",
        "joppelicious": "Johannes",
        "KaffeKaffe": "Fredrik",
        "Kamomillan": "Camilla",
        "kim.stenberg": "Kim",
        "landfelt": "Niklas",
        "liljekvist02": "MikaelL", # This Mikael has been modified to MikaelL to not collide with Migus-Mikael
        "Migus": "Mikael",
        "ministerkrister": "Kristofer",
        "peterjaric": "Peter",
        "proso": "PelleA",
        "Scrbell": "Jöran",
        "snilser93": "Svante",
        "ulrika.b.carlsson": "Ulrika",
        "uumartin": "Martin",
        "Åkermarken": "Lars",
    }

    # Loop through all sheets
    print("Looping through sheets")
    for sheet in workbook.worksheets:
        print('Handling ' + sheet.title)
        realname_to_wordfeudname = {}

        # Only parse sheets that are seasons or cups
        if (sheet.title.startswith("Säsong") or sheet.title.startswith("Cup")):
            # Parse player names and Wordfeud aliases
            if sheet.title == "Säsong 1":
                # The first season doesn't list the players' Wordfeud names
                realname_to_wordfeudname = {
                    "Jonas": "drsadness",
                    "Pelle": "Äggeth",
                    "Per": "gurka1495",
                    "Kristofer": "ministerkrister"
                }
            else:
                # Find the cell with the name "Wordfeudnamn"
                wordfeud_name_columns = []
                wordfeud_name_rows = []
                for row in sheet.iter_rows(min_row=1, max_row=100, min_col=1, max_col=30):
                    for cell in row:
                        if isinstance(cell.value, str) and cell.value.strip() == "Wordfeudnamn":
                            wordfeud_name_columns.append(cell.column)
                            wordfeud_name_rows.append(cell.row)


                for index, wordfeud_name_column in enumerate(wordfeud_name_columns):
                    wordfeud_name_row = wordfeud_name_rows[index]
                    # Connect the player names with their Wordfeud aliases by iterating through the rows under the cell with the name "Wordfeudnamn"
                    for row in sheet.iter_rows(min_row=wordfeud_name_row + 1, min_col=wordfeud_name_column - 1, max_col=wordfeud_name_column):
                        if row[0].value:
                            realname_to_wordfeudname[row[0].value.strip()] = row[1].value.strip()
                        else:
                            # Loop until we find an empty cell
                            break
            # Parse results
            current_matches = []

            # Find sets of cells that contain match results
            for row in sheet.iter_rows(min_row=2, min_col=1, max_col=20):
                for i in range(0, len(row) - 4):
                    try:
                        value0 = row[i].value
                        value1 = row[i + 1].value
                        value2 = row[i + 2].value
                        value3 = row[i + 3].value
                        value4 = row[i + 4].value
                        value5 = row[i + 5].value if len(row) > i + 6 else None

                        if is_season_score(value0, value1, value2, value4, value5):
                            player_1_name = value0
                            player_2_name = value1
                            player_1_score = value4
                            player_2_score = value5
                        elif is_cup_score(value0, value1, value2, value3, value4):
                            player_1_name = value0
                            player_2_name = value1
                            player_1_score = value2
                            player_2_score = value3
                        elif is_alternate_cup_score(value0, value1, value2, value3, value4):
                            player_1_name = value0
                            player_2_name = value1
                            player_1_score = value3
                            player_2_score = value4
                        else:
                            continue

                        match = {
                            "sheet": sheet.title,
                            "player1": wordfeudname_to_realname[realname_to_wordfeudname[player_1_name.strip()]],
                            "player2": wordfeudname_to_realname[realname_to_wordfeudname[player_2_name.strip()]],
                            "score1": int(player_1_score),
                            "score2": int(player_2_score)
                        }
                        current_matches.append(match)
                    except Exception as e:
                        log.exception("Error parsing match")

            # Add matches at the start of the list since sheets are listed in reverse order
            matches = current_matches + matches

            if len(current_matches) == 0:
                log.error('No matches in ' + sheet.title)

    # Write the matches to the file
    # convert start date to seconds since epoch
    start_time = time.mktime(time.strptime("2023-01-01", "%Y-%m-%d"))

    print("Writing to " + rayter_file)

    for match in matches:
        # Convert seconds since epoch to date
        date = time.strftime("%Y-%m-%d %H:%M", time.localtime(start_time))
        f.write("game " + date + "\n")
        f.write(match["player1"] + "\t\t" + str(match["score1"]) + "\n")
        f.write(match["player2"] + "\t\t" + str(match["score2"]) + "\n")
        f.write("\n")

        # Increase the date by 1 day
        start_time += 24 * 60 * 60
